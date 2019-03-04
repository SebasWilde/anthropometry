from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import get_column_letter, column_index_from_string


def style_range(ws, cell_range, style=None):
    """
    :param ws:  Excel worksheet instance
    :param cell_range: An excel range to style (e.g. A1:F20)
    :param style: An openpyxl Style object
    """

    start_cell, end_cell = cell_range.split(':')
    start_coord = coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = column_index_from_string(start_coord[0])
    end_coord = coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(column=col, row=row).style = style


def merge_with_right(ws, cell_range, number__merge_cells):
    start_cell, end_cell = cell_range.split(':')
    start_coord = coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = column_index_from_string(start_coord[0])
    end_coord = coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = column_index_from_string(end_coord[0])
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1, number__merge_cells):
            init_col = get_column_letter(col)
            next_col = get_column_letter(col+1)
            ws.merge_cells('{0}{2}:{1}{2}'.format(init_col, next_col, row))
