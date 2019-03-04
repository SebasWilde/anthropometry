from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse_lazy

from .models import Deportista, Medida
from .constants import DECIMAL_FORMAT, PERCENTAGE_FORMAT, DATE_STYLE, ONE_DECIMAL_STYLE, MEDIUM_BORDER, MEDIUM_BORDER_STYLE, DEPORTISTA_STYLE, RESULTS_TITLE_STYLE, ALIGN_MIDDLE, THIN_BORDER
from .utils import style_range, merge_with_right
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Side, Border, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import (
    Reference,
    BarChart
)


class ReporterTrainerExcel(LoginRequiredMixin, View):

    def set_dimensions(self, ws):
        ws.column_dimensions['A'].width = 3.5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 13
        for col in range(4, column_index_from_string('Y')+1):
            ws.column_dimensions[get_column_letter(col)].width = 8
        ws.column_dimensions['Z'].width = 20
        ws.column_dimensions['AA'].width = 40
        ws.row_dimensions[7].height = 25

    def set_table(self, ws, row_next, mediciones):
        # Header table
        font_header = Font(name='Tw Cen MT Condensed', size=10, bold=True)
        align_header = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        border_special_head = Border(left=Side(style='medium'),
                                     right=Side(style='medium'),
                                     top=Side(style='medium'),
                                     bottom=Side(style='thin'))

        border_special_head_2 = Border(left=Side(style='thin'),
                                       right=Side(style='thin'),
                                       top=Side(style='thin'),
                                       bottom=Side(style='medium'))

        border_special_head_3 = Border(left=Side(style='thin'),
                                       right=Side(style='medium'),
                                       top=Side(style='thin'),
                                       bottom=Side(style='medium'))

        header = ['N°', 'Apellidos y Nombres', 'F. de Eval', 'Edad',
                  'Peso bruto (kg)', 'Estatura (cm)', 'IMC', 'Talla//Edad',
                  'IMC/Edad', 'Masa magra (kg)',
                  'Sum pliegues (mm)', '% Grasa corporal', 'Talla sentado',
                  'Indice citura cadera', 'TRC', 'SUBS', 'BICP', 'SPRI',
                  'SPRE', 'ABD', 'MSM', 'PANT', 'Endo', 'Meso', 'Ecto',
                  'Catacterización', 'Observación']

        for col in range(0, column_index_from_string('G')):
            column = get_column_letter(col+1)
            range_columns = '{0}{1}:{0}{2}'.format(column, row_next-1, row_next)
            ws.merge_cells(range_columns)
            style_range(ws, range_columns, MEDIUM_BORDER_STYLE)

        for col in range(column_index_from_string('J'),
                         column_index_from_string('O')):
            column = get_column_letter(col)
            range_columns = '{0}{1}:{0}{2}'.format(column, row_next-1, row_next)
            ws.merge_cells(range_columns)
            style_range(ws, range_columns, MEDIUM_BORDER_STYLE)

        ws.merge_cells('AA{0}:AA{1}'.format(row_next - 1, row_next))
        style_range(ws, 'AA{0}:AA{1}'.format(row_next - 1, row_next),
                    MEDIUM_BORDER_STYLE)

        ws.merge_cells('H6:I6')
        head = ws.cell(row=6, column=column_index_from_string('H'), value='SALUD')
        head.font = font_header
        head.alignment = align_header
        head.border = border_special_head

        ws.merge_cells('O6:V6')
        head = ws.cell(row=6, column=column_index_from_string('O'),
                       value='Pliegues Cutáneos (mm)')
        head.font = font_header
        head.alignment = align_header
        head.border = border_special_head

        ws.merge_cells('W6:Z6')
        head = ws.cell(row=6, column=column_index_from_string('W'),
                       value='Somatotipo')
        head.font = font_header
        head.alignment = align_header
        head.border = border_special_head

        second_data = ['Talla//Edad', 'IMC/Edad', 'TRC', 'SUBS', 'BICP', 'SPRI',
                       'SPRE', 'ABD', 'MSM', 'PANT', 'Endo', 'Meso', 'Ecto',
                       'Catacterización']

        for col in range(0, len(header)):
            if header[col] in second_data:
                head = ws.cell(column=col + 1, row=row_next,
                               value='{0}'.format(header[col]))
                if header[col] == 'IMC/Edad' or header[col] == 'PANT':
                    head.border = border_special_head_3
                else:
                    head.border = border_special_head_2

            else:
                head = ws.cell(column=col + 1, row=row_next-1,
                               value='{0}'.format(header[col]))
            head.font = font_header
            head.alignment = align_header

        row_next += 1

        # Table content

        date_style = NamedStyle(name='date', number_format='DD/MM/YYYY')
        decimal_one_style = NamedStyle(name='decimal', number_format='0.0')
        font_table_content = Font(name='Tw Cen MT Condensed Extra Bold',
                                  size=11)

        border_both_separtor = Border(left=Side(style='medium'),
                                      right=Side(style='medium'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))

        border_left_separtor = Border(left=Side(style='medium'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))

        border_right_separtor = Border(left=Side(style='thin'),
                                       right=Side(style='medium'),
                                       top=Side(style='thin'),
                                       bottom=Side(style='thin'))

        for row in range(0, mediciones.count()):
            col = 1
            data = ws.cell(column=col, row=row_next, value=row + 1)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_both_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value='{0}'.format(mediciones[row]
                                              .deportista.get_full_name()))
            data.font = font_table_content
            data.border = border_both_separtor
            col += 1

            fecha_registro = ws.cell(column=col, row=row_next,
                                     value=mediciones[row].fecha_registro)
            fecha_registro.style = date_style
            fecha_registro.font = font_table_content
            fecha_registro.alignment = ALIGN_MIDDLE
            fecha_registro.border = border_both_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].get_age_at_moment_metrics())
            data.font = font_table_content
            data.border = border_both_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].peso_bruto)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_both_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].estatura)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_both_separtor
            col += 1

            imc = ws.cell(column=col, row=row_next,
                          value=mediciones[row].get_indice_masa_corporal())
            imc.style = decimal_one_style
            imc.font = font_table_content
            imc.alignment = ALIGN_MIDDLE
            imc.border = border_both_separtor
            col += 3

            masa_magra = ws.cell(column=col, row=row_next,
                                 value=mediciones[row].get_masa_magra())
            masa_magra.style = decimal_one_style
            masa_magra.font = font_table_content
            masa_magra.alignment = ALIGN_MIDDLE
            masa_magra.border = border_both_separtor
            col += 1

            s_pligues = ws.cell(column=col, row=row_next,
                                value=mediciones[row].get_sum_pliegues())
            s_pligues.style = decimal_one_style
            s_pligues.font = font_table_content
            s_pligues.alignment = ALIGN_MIDDLE
            s_pligues.border = border_both_separtor
            col += 1

            grasa_corporal = ws.cell(column=col, row=row_next,
                                     value=mediciones[
                                         row].get_grasa_corporal())
            grasa_corporal.number_format = PERCENTAGE_FORMAT
            grasa_corporal.font = font_table_content
            grasa_corporal.alignment = ALIGN_MIDDLE
            grasa_corporal.border = border_both_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].talla_sentado)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_both_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].get_indice_cintaura_cadera())
            data.style = decimal_one_style
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_both_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].triceps)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_left_separtor
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].subescapular)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = THIN_BORDER
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].biceps)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = THIN_BORDER
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].suprailiaco)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = THIN_BORDER
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].supraespinal)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = THIN_BORDER
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].abdominales)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = THIN_BORDER
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].muslo_medio)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = THIN_BORDER
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value=mediciones[row].pliege_pierna)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_right_separtor
            col += 1

            endo = ws.cell(column=col, row=row_next,
                           value=mediciones[row].get_endomorfismo())
            endo.style = decimal_one_style
            endo.font = font_table_content
            endo.alignment = ALIGN_MIDDLE
            endo.border = border_left_separtor
            col += 1

            meso = ws.cell(column=col, row=row_next,
                           value=mediciones[row].get_mesomorfismo())
            meso.style = decimal_one_style
            meso.font = font_table_content
            meso.alignment = ALIGN_MIDDLE
            meso.border = THIN_BORDER
            col += 1

            ecto = ws.cell(column=col, row=row_next,
                           value=mediciones[row].get_ectomorfismo())
            ecto.style = decimal_one_style
            ecto.font = font_table_content
            ecto.alignment = ALIGN_MIDDLE
            ecto.border = THIN_BORDER
            col += 1

            data = ws.cell(column=col, row=row_next,
                           value='{0}'.format(
                               mediciones[row].get_caracterizacion()))
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_right_separtor
            col += 1

            data = ws.cell(column=col,
                           row=row_next,
                           value='{0}'
                           .format(mediciones[row].get_observacion_masa_grasa()[1]
                                   + '/' +
                                   mediciones[
                                       row].get_observacion_masa_muscular()[1]))
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.border = border_both_separtor
            col += 1
            row_next += 1

        style_border_left = NamedStyle('style_border_left_table',
                                       font=font_table_content,
                                       alignment=ALIGN_MIDDLE,
                                       border=border_left_separtor)
        style_border_right = NamedStyle('style_border_right_table',
                                        font=font_table_content,
                                        alignment=ALIGN_MIDDLE,
                                        border=border_right_separtor
                                        )
        style_range(ws, 'H8:H{0}'.format(row_next-1), style_border_left)
        style_range(ws, 'I8:I{0}'.format(row_next-1), style_border_right)

        col_next = len(header) + 1

        observacion = ws.cell(column=2, row=row_next+2, value='Observación:')
        observacion.font = Font(name='Arial', size=10, bold=True)
        text = 'Todos los nombres marcados con verde son deportistas ' \
               'catalogados como "madurador temprano" es decir son ' \
               'aquellos que ya han tenido su pico de maduración máximo y es ' \
               'probable que puedan tolerar el entrenamiento mejor que el ' \
               'resto, tener mejores tiempo de recuperación, mayor fuerza, ' \
               'resistencia, etc. Los deportistas en blanco son "maduradores ' \
               'normales" y probablemente tengan su pico de maduración en el ' \
               'transcurso del año. Finalmente, todos los marcados con ' \
               'amarillo son "maduradores tardíos" pero probablemente su pico ' \
               'se dará en más de un año. El indicador Talla//Edad nos ' \
               'muestra a los deportistas que presentan la relación de la ' \
               'talla para la edad y según esto son catalogados como ' \
               '"Altos", "Muy altos" o "Normales"'

        ws.merge_cells('F{0}:Y{0}'.format(row_next+2))
        observacion_text = ws.cell(column=6, row=row_next+2, value=text)
        observacion_text.font = Font(name='Arial', size=10)
        lic = ws.cell(column=col_next-1, row=row_next+2,
                      value='Lic. José Miguel Chambi Enríquez')
        lic.font = Font(name='Arial', size=10, bold=True)
        cnp = ws.cell(column=col_next - 1, row=row_next + 3,
                      value='CNP : 5582')
        cnp.font = Font(name='Arial', size=10, bold=True)
        row_next += 4

        return row_next, col_next

    def set_header(self, ws, col_next, institucion, categoria):
        categoria_label = ws.cell(column=col_next - 1, row=1,
                                  value='{0}'.format(institucion).upper())
        categoria_label.font = Font(name='Arial', size=9, bold=True)
        categoria_label.alignment = Alignment(horizontal='right')
        area_label = ws.cell(column=col_next - 1, row=2,
                             value='AREA DE NUTRICIÓN ')
        area_label.font = Font(name='Arial', size=9, bold=True)
        area_label.alignment = Alignment(horizontal='right')

        ws.merge_cells('A3:' + get_column_letter(col_next-1)+str(3))
        title = ws.cell(column=1, row=3, value='INFORME DE EVALUACIÓN '
                                               'ANTROPOMÉTRICA - {0}'.
                        format(categoria).upper())
        title.font = Font(name='Arial', size=14, bold=True)
        title.alignment = ALIGN_MIDDLE

    def get(self, request, *args, **kwargs):
        deporte = request.GET.get('deporte', None)
        date_input = request.GET.get('date', None)
        if not deporte:
            return HttpResponseRedirect(reverse_lazy('reporter'))
        categoria = request.GET.get('categoria', None)
        if categoria:
            deportistas = Deportista.objects.filter(deporte=deporte,
                                                    categoria=categoria)
        else:
            deportistas = Deportista.objects.filter(deporte=deporte)

        if date_input:
            year, month = date_input.split('-')
            mediciones = Medida.objects.filter(
                deportista__in=deportistas,
                fecha_registro__month__gte=month,
                fecha_registro__year__gte=year
            ).order_by('-fecha_registro')
        else:
            mediciones = Medida.objects.filter(deportista__in=deportistas) \
                .order_by('-fecha_registro')

        institucion = mediciones.first().deportista.institucion
        categoria = mediciones.first().deportista.categoria

        wb = Workbook()
        ws = wb.active

        row_next = 7
        row_next, col_next = self.set_table(ws, row_next, mediciones)
        self.set_header(ws, col_next, institucion, categoria)
        self.set_dimensions(ws)
        file_name = '{0}_{1}_{2}'.format(institucion,
                                         mediciones.first().deportista.deporte,
                                         mediciones.first().deportista.categoria)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{0}.xls"' \
            .format(file_name)
        wb.save(response)
        return response


class ReporterDeportistaExcel(LoginRequiredMixin, View):

    def set_header(self, ws):
        ws.merge_cells('G1:M1')
        head = ws.cell(column=column_index_from_string('G'), row=1)
        head.value = 'ÁREA DE NUTRICIÓN'
        head.font = Font(name='Gill Sans MT', bold=True, size=10)
        head.alignment = Alignment(horizontal='center')
        ws.merge_cells('B3:R3')
        title = ws.cell(column=column_index_from_string('B'), row=3)
        title.value = 'FICHA RESUMEN DE EVALUACIÓN ANTROPOMÉTRICA'
        title.font = Font(name='Arial Black', bold=True, size=14)
        title.alignment = Alignment(horizontal='center')

    def set_deportista_section(self, ws, deportista):
        ws.merge_cells('B5:C5')
        deportista_label = ws.cell(column=2, row=5)
        deportista_label.value = 'Deportista:'
        deportista_label.style = DEPORTISTA_STYLE

        ws.merge_cells('D5:M5')
        deportista_name = ws.cell(column=column_index_from_string('D'), row=5)
        deportista_name.value = deportista.__str__()
        style_range(ws, 'D5:M5', MEDIUM_BORDER_STYLE)
        deportista_name.font = Font(name='Calibri', bold=True, size=11)

        ws.merge_cells('B7:C7')
        nacimiento_label = ws.cell(column=2, row=7)
        nacimiento_label.value = 'F. de Nac:'
        nacimiento_label.style = DEPORTISTA_STYLE

        ws.merge_cells('D7:G7')
        nacimiento = ws.cell(column=column_index_from_string('D'), row=7)
        style_range(ws, 'D7:G7', MEDIUM_BORDER_STYLE)
        nacimiento.value = deportista.fecha_nacimiento
        nacimiento.number_format = 'DD/MM/YYYY'
        nacimiento.font = Font(name='Calibri', bold=True, size=11)
        nacimiento.alignment = Alignment(horizontal='center')

        ws.merge_cells('H7:I7')
        posicion_label = ws.cell(column=column_index_from_string('H'), row=7)
        posicion_label.value = 'Posición'
        posicion_label.style = DEPORTISTA_STYLE

        ws.merge_cells('N7:O7')
        categoria_label = ws.cell(column=column_index_from_string('N'), row=7)
        categoria_label.value = 'Categoría:'
        categoria_label.style = DEPORTISTA_STYLE

        ws.merge_cells('P7:Q7')
        categoria_data = ws.cell(column=column_index_from_string('P'), row=7)
        categoria_data.value = deportista.categoria.__str__()
        style_range(ws, 'P7:Q7', MEDIUM_BORDER_STYLE)
        categoria_data.font = Font(name='Calibri', bold=True, size=11)
        categoria_data.alignment = Alignment(horizontal='center')

    def set_table_results(self, ws, mediciones, font_header_table,
                          font_table_content):
        ws.merge_cells('B9:R9')
        ws.cell(column=2, row=9,
                value='RESULTADOS').style = RESULTS_TITLE_STYLE

        # Set styles in table results

        merge_with_right(ws, 'C11:P13', 2)
        style_range(ws, 'C11:P13', MEDIUM_BORDER_STYLE)

        merge_with_right(ws, 'C29:P29', 2)
        style_range(ws, 'C29:P29', MEDIUM_BORDER_STYLE)

        merge_with_right(ws, 'C45:P46', 2)
        style_range(ws, 'C45:P46', MEDIUM_BORDER_STYLE)

        merge_with_right(ws, 'C48:P51', 2)
        style_range(ws, 'C48:P51', MEDIUM_BORDER_STYLE)

        # Set data in table results
        label = ws.cell(row=11, column=column_index_from_string('C'),
                        value='Fecha')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=12, column=column_index_from_string('C'),
                        value='Edad')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=13, column=column_index_from_string('C'),
                        value='Estatura')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=29, column=column_index_from_string('C'),
                        value='Peso')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=45, column=column_index_from_string('C'),
                        value='% Grasa(YC)')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=46, column=column_index_from_string('C'),
                        value='\u03A3 6 Pliegues')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=48, column=column_index_from_string('C'),
                        value='IMC')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=49, column=column_index_from_string('C'),
                        value='AKS')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=50, column=column_index_from_string('C'),
                        value='Talla//Edad')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(row=51, column=column_index_from_string('C'),
                        value='IMC//Edad')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        start_column = column_index_from_string('E')

        for medida in mediciones:
            data = ws.cell(row=11, column=start_column,
                           value=medida.fecha_registro.strftime('%d-%b-%y'))
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.number_format = 'DD-mm-yy'

            data = ws.cell(row=12, column=start_column,
                           value=medida.get_age_at_moment_metrics())
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE

            data = ws.cell(row=13, column=start_column, value=medida.estatura)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE

            data = ws.cell(row=29, column=start_column,
                           value=medida.peso_bruto)
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE

            data = ws.cell(row=45, column=start_column,
                           value=medida.get_grasa_corporal())
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.number_format = PERCENTAGE_FORMAT

            data = ws.cell(row=46, column=start_column,
                           value=medida.get_sum_pliegues())
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.number_format = DECIMAL_FORMAT

            data = ws.cell(row=48, column=start_column,
                           value=medida.get_indice_masa_corporal())
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.number_format = DECIMAL_FORMAT

            data = ws.cell(row=49, column=start_column,
                           value=medida.get_indice_sustancia_activa())
            data.font = font_table_content
            data.alignment = ALIGN_MIDDLE
            data.number_format = DECIMAL_FORMAT

            start_column += 2

    def set_categorizacion(self, ws, mediciones, font_header_table,
                           font_table_content):
        ws.merge_cells('B53:R53')
        ws.cell(column=2, row=53, value='CATEGORIZACIÓN') \
            .style = RESULTS_TITLE_STYLE

        ws.merge_cells('B55:C57')
        style_range(ws, 'B55:C57', MEDIUM_BORDER_STYLE)
        label = ws.cell(column=2, row=55, value='Masa Grasa')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        ws.merge_cells('B59:C62')
        style_range(ws, 'B59:C62', MEDIUM_BORDER_STYLE)
        label = ws.cell(column=2, row=59, value='Masa Muscular')
        label.font = font_header_table
        label.alignment = ALIGN_MIDDLE

        merge_with_right(ws, 'D55:Q57', 2)
        style_range(ws, 'D55:Q57', MEDIUM_BORDER_STYLE)
        merge_with_right(ws, 'D59:Q62', 2)
        style_range(ws, 'D59:Q62', MEDIUM_BORDER_STYLE)

        label = ws.cell(column=4, row=55, value='Adecuada')
        label.font = Font(name='Calibri', size=8, bold=True)
        label.alignment = ALIGN_MIDDLE
        label = ws.cell(column=4, row=56, value='Lig Elevada')
        label.font = Font(name='Calibri', size=8, bold=True)
        label.alignment = ALIGN_MIDDLE
        label = ws.cell(column=4, row=57, value='Elevada')
        label.font = Font(name='Calibri', size=8, bold=True)
        label.alignment = ALIGN_MIDDLE

        label = ws.cell(column=4, row=59, value='Muy escasa')
        label.font = Font(name='Calibri', size=8, bold=True)
        label.alignment = ALIGN_MIDDLE
        label = ws.cell(column=4, row=60, value='Escasa')
        label.font = Font(name='Calibri', size=8, bold=True)
        label.alignment = ALIGN_MIDDLE
        label = ws.cell(column=4, row=61, value='Adecuada')
        label.font = Font(name='Calibri', size=8, bold=True)
        label.alignment = ALIGN_MIDDLE
        label = ws.cell(column=4, row=62, value='Elevada')
        label.font = Font(name='Calibri', size=8, bold=True)
        label.alignment = ALIGN_MIDDLE

        start_column = column_index_from_string('F')
        for medida in mediciones:
            masa_grasa = medida.get_observacion_masa_grasa()[1]
            if 'Muy elevada' in masa_grasa:
                data = ws.cell(row=57, column=start_column, value='X')
                data.font = font_table_content
                data.alignment = ALIGN_MIDDLE
            elif 'Elevada' in masa_grasa:
                data = ws.cell(row=56, column=start_column, value='X')
                data.font = font_table_content
                data.alignment = ALIGN_MIDDLE
            elif 'Adecuada' in masa_grasa:
                data = ws.cell(row=55, column=start_column, value='X')
                data.font = font_table_content
                data.alignment = ALIGN_MIDDLE

            masa_muscular = medida.get_observacion_masa_muscular()[1]
            if 'Muy escasa' in masa_muscular:
                data = ws.cell(row=59, column=start_column, value='X')
                data.font = font_table_content
                data.alignment = ALIGN_MIDDLE
            elif 'Escasa' in masa_muscular:
                data = ws.cell(row=60, column=start_column, value='X')
                data.font = font_table_content
                data.alignment = ALIGN_MIDDLE
            elif 'Adecuada' in masa_muscular:
                data = ws.cell(row=61, column=start_column, value='X')
                data.font = font_table_content
                data.alignment = ALIGN_MIDDLE
            elif 'Elevada' in masa_muscular:
                data = ws.cell(row=62, column=start_column, value='X')
                data.font = font_table_content
                data.alignment = ALIGN_MIDDLE
            start_column += 2

    def set_charts(self, ws):
        chart_peso = BarChart()
        chart_peso.type = "col"
        chart_peso.style = 10
        chart_peso.title = "Peso"
        data = Reference(ws, min_col=5, min_row=29, max_col=15)
        chart_peso.add_data(data, from_rows=True)
        chart_peso.shape = 4
        ws.add_chart(chart_peso, "C14")

        chart_grasa = BarChart()
        chart_grasa.type = "col"
        chart_grasa.style = 10
        chart_grasa.title = "%GRASA (YC)"
        data = Reference(ws, min_col=5, min_row=45, max_col=15)
        chart_grasa.add_data(data, from_rows=True)
        chart_grasa.shape = 4
        ws.add_chart(chart_grasa, "C30")

    def get(self, request, *args, **kwargs):
        deportista_pk = request.GET.get('deportista', None)
        date_input = request.GET.get('date', None)
        if not deportista_pk:
            return HttpResponseRedirect(reverse_lazy('reporter'))
        deportista = Deportista.objects.get(pk=deportista_pk)
        if date_input:
            year, month = date_input.split('-')
            mediciones = Medida.objects.filter(
                deportista=deportista,
                fecha_registro__month__gte=month,
                fecha_registro__year__gte=year
            ).order_by('fecha_registro')[:6]
        else:
            mediciones = Medida.objects.filter(deportista=deportista) \
                             .order_by('fecha_registro')[:6]

        wb = Workbook()
        ws = wb.active

        # Header section
        self.set_header(ws)

        font_header_table = font = Font(name='Calibri', size=10, bold=True)
        font_table_content = Font(name='Calibri', size=11)

        # Deportista section
        self.set_deportista_section(ws, deportista)

        # Resultados section
        self.set_table_results(ws, mediciones, font_header_table,
                               font_table_content)

        # Section categorazacion
        self.set_categorizacion(ws, mediciones, font_header_table,
                                font_table_content)
        # Set charts
        self.set_charts(ws)

        for col in range(1, column_index_from_string('Q')+1):
            ws.column_dimensions[get_column_letter(col)].width = 6

        response = HttpResponse(content_type='application/ms-excel')
        filename = '{0}.xlsx'.format(deportista.get_full_name())
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        wb.save(response)
        return response
